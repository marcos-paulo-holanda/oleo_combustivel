'''
https://olinda.bcb.gov.br/olinda/servico/PTAX/versao/v1/aplicacao#!/recursos/CotacaoDolarPeriodo

Captura em tempo real o valor de cotação do Petróleo Brent
Link de acesso: https://oilprice.com/oil-price-charts/#Brent-Crude
'''
import requests as req, numpy as np, pandas as pd
from datetime import datetime, timedelta
from bs4 import BeautifulSoup
import openpyxl as op

def dolar():
        
    dI = datetime.today().strftime("%m-%d-%Y")
    dF = datetime.today() - timedelta(days=30)
    dF = dF.strftime("%m-%d-%Y")

    prd = "https://olinda.bcb.gov.br/olinda/servico/PTAX/versao/v1/odata/CotacaoDolarPeriodo(dataInicial=@dataInicial,dataFinalCotacao=@dataFinalCotacao)?@dataInicial=" \
    + "'" + dF + "'" + "&@dataFinalCotacao=" + "'" + dI + "'" + "&$top=100000&$format=json&$select=cotacaoCompra,cotacaoVenda,dataHoraCotacao"

    cotacao = req.get(prd).json()

    datas, venda = [], []

    for i in range(len(cotacao['value'])):
        venda.append(cotacao['value'][i]['cotacaoVenda'])
        datas.append(cotacao['value'][i]['dataHoraCotacao'])

    datas = [x[:10] for x in datas]

    #df = pd.DataFrame(list(zip(datas,venda)), columns =['Data', 'Valor de venda'])

    return datas, venda

#df.to_excel(r'dolarCotacao.xlsx', sheet_name='dolar', index=False)


def brent():
    url = 'https://oilprice.com/oil-price-charts/#Brent-Crude'

    r = req.get(url)
    html_doc = r.text
    soup = BeautifulSoup(html_doc, "html.parser")

    tag = soup.find_all('td', {'class':'last_price'})

    brent_last_price = tag[1].get_text()

    return brent_last_price

def write_excel(datas, vendas, brent, nome_tabela):
    wb = op.load_workbook(nome_tabela + '.xlsx')
    st = wb.worksheets[0]
    for i in range(len(datas)):
        st.cell(row = i+2, column = 2).value = datas[i]
        st.cell(row = i+2, column = 3).value = vendas[i]
        
    wb.save(nome_tabela + '.xlsx')

datas, venda = dolar()
b = brent()
write_excel(datas, venda, b, 'fuel_oil')
